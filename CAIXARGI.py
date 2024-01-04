from tkinter import Frame
from openpyxl import Workbook, load_workbook
import PySimpleGUI as sg

tema = "SystemDefault"

#funções

def Escolha():
        #layout
        sg.theme(tema)
        
        layout = [
            [sg.Text("Funcionário: "), sg.Combo(["Renata", "Suellen", "Moises", "Endrielly"])],
            [sg.Text("Dia:"), sg.Combo(["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12", "13", "14", "15", "16", "17", "18", "19", "20", "21", "22", "23", "24", "25", "26", "27", "28", "29", "30", "31"])],
            [sg.Button("ABRIR CAIXA", size=(20,1))]
        ]

        #janela
        janela = sg.Window("CAIXA RGI").layout(layout)

        event, values = janela.read()

        fun = values[0]
        dia = values[1]
        
        janela.Close()
        return fun, dia
def Entrada():

        #layout
        sg.theme(tema)

        
        layout = [
            [sg.Text('TROCO INICIAL: ', size=(20,1)), sg.Input(size=(10,1), key="TrocoInicial")],
            [sg.Text('ENTRADAS: ', size=(20,1)), sg.Input(size=(10,1), key="Entradas")],
            [sg.Text("", size=(20,1))],
            [sg.Text('1 - CARTÃO DE DÉBITO: ', size=(20,1)), sg.Input(size=(10,1), key="CartDEB")],
            [sg.Text('2 - DEPÓSITO BANCÁRIO: ', size=(20,1)), sg.Input(size=(10,1), key="DepoBAN")],
            [sg.Text('3 - CHEQUE PRÉ: ', size=(20,1)), sg.Input(size=(10,1), key="ChequePRE")],
            [sg.Text('4 - CHEQUE DIA: ', size=(20,1)), sg.Input(size=(10,1), key="ChequeDIA")],
            [sg.Text('5 - DEPÓSITO INTERNO: ', size=(20,1)), sg.Input(size=(10,1), key="DepositoINTERNO")],
            [sg.Text('6 - OFÍCIO ELETRÔNICO: ', size=(20,1)), sg.Input(size=(10,1), key="OficioELE"), sg.Text('Doc nº: ', size=(7,1)), sg.Input(size=(30,1), key="doc1")],
            [sg.Text('7 - E-INTIMAÇÃO: ', size=(20,1)), sg.Input(size=(10,1), key="Einti"), sg.Text('Doc nº: ', size=(7,1)), sg.Input(size=(30,1), key="doc2")], 
            [sg.Text('8 - ERRO DE FUNCIONÁRIO: ', size=(20,1)), sg.Input(size=(10,1), key="ErroFun"), sg.Text('Doc nº: ', size=(7,1)), sg.Input(size=(30,1), key="doc3")],
            [sg.Text('9 - ERRO DO SISTEMA: ', size=(20,1)), sg.Input(size=(10,1), key="ErroSis"), sg.Text('Doc nº: ', size=(7,1)), sg.Input(size=(30,1), key="doc4")],
            [sg.Text('10 - DMP: ', size=(20,1)), sg.Input(size=(10,1), key="dmp"), sg.Text('Doc nº: ', size=(7,1)), sg.Input(size=(30,1), key="doc5"), sg.Text('Empresa: ', size=(7,1)), sg.Input(size=(30,1), key="empresa")], 
            [sg.Text('11 - DÉBITO AUTORIZADO: ', size=(20,1)), sg.Input(size=(10,1), key="DebitoAuto"), sg.Text('Doc nº: ', size=(7,1)), sg.Input(size=(30,1), key="doc6"), sg.Text('Por: ', size=(7,1)), sg.Input(size=(30,1), key="por1")],
            [sg.Text('12 - NIHIL: ', size=(20,1)), sg.Input(size=(10,1), key="Nihil"), sg.Text('Doc nº: ', size=(7,1)), sg.Input(size=(30,1), key="doc7"), sg.Text('Por: ', size=(7,1)), sg.Input(size=(30,1), key="por2")],
            [sg.Text('Quantidades de Transf:', size=(20,1)), sg.Input(size=(10,1), key="TransfQtT"), sg.Text('Espécie: ', size=(10,1)), sg.Input(size=(10,1), key="TransfEspe"), sg.Text('Cheque Pré: ', size=(10,1)), sg.Input(size=(10,1), key="TransfPre"), sg.Text('Cheque Dia: ', size=(10,1)), sg.Input(size=(10,1), key="TransfDia")],
            [sg.Text("", size=(20,1))],
            [sg.Text('CAIXA FINAL: ', size=(20,1)), sg.Input(size=(10,1), key="CaixaFinal")],
            [sg.Button("Enviar", size=(20,1))], 
        ]

        #janela
        janela = sg.Window("CAIXA").layout(layout)

        event, values = janela.read()

        if values["TrocoInicial"] == "":
            values["TrocoInicial"] = "0"
        
        if values["Entradas"] == "":
            values["Entradas"] = "0"  
        
        if values["DepoBAN"] == "":
            values["DepoBAN"] = "0" 

        if values["ChequePRE"] == "":
            values["ChequePRE"] = "0"  

        if values["ChequeDIA"] == "":
            values["ChequeDIA"] = "0"  

        if values["DepositoINTERNO"] == "":
            values["DepositoINTERNO"] = "0"  

        if values["OficioELE"] == "":
            values["OficioELE"] = "0"  

        if values["Einti"] == "":
            values["Einti"] = "0"  

        if values["ErroFun"] == "":
            values["ErroFun"] = "0"  
        
        if values["ErroSis"] == "":
            values["ErroSis"] = "0" 

        if values["dmp"] == "":
            values["dmp"] = "0"  

        if values["DebitoAuto"] == "":
            values["DebitoAuto"] = "0"  

        if values["Nihil"] == "":
            values["Nihil"] = "0" 

        if values["CaixaFinal"] == "":
            values["CaixaFinal"] = "0"  

        if values["CartDEB"] == "":
            values["CartDEB"] = "0"  

        if values["TransfQtT"] == "":
            values["TransfQtT"] = "0" 

        if values["TransfEspe"] == "":
            values["TransfEspe"] = "0" 

        if values["TransfPre"] == "":
            values["TransfPre"] = "0" 

        if values["TransfDia"] == "":
            values["TransfDia"] = "0" 

        TrocoInicial = values["TrocoInicial"].replace(",", ".")
        Entradas = values["Entradas"].replace(",", ".")
        CartDEB = values["CartDEB"].replace(",", ".")
        DepoBAN = values["DepoBAN"].replace(",", ".")
        ChequePRE = values["ChequePRE"].replace(",", ".")
        ChequeDIA = values["ChequeDIA"].replace(",", ".")
        DepositoINTERNO = values["DepositoINTERNO"].replace(",", ".")
        OficioELE = values["OficioELE"].replace(",", ".")
        Einti = values["Einti"].replace(",", ".")
        ErroFun = values["ErroFun"].replace(",", ".")
        ErroSis = values["ErroSis"].replace(",", ".")
        dmp = values["dmp"].replace(",", ".") 
        DebitoAuto = values["DebitoAuto"].replace(",", ".")
        Nihil = values["Nihil"].replace(",", ".")
        CaixaFinal = values["CaixaFinal"].replace(",", ".")
        xdoc1 = values["doc1"]
        xdoc2 = values["doc2"]
        xdoc3 = values["doc3"]
        xdoc4 = values["doc4"]
        xdoc5 = values["doc5"]
        xdoc6 = values["doc6"]
        xdoc7 = values["doc7"]
        xempresa = values["empresa"]
        xpor1 = values["por1"]
        xpor2 = values["por2"]
        TransfQtT = values["TransfQtT"].replace(",", ".")
        TransfEspe = values["TransfEspe"].replace(",", ".")
        TransfPre = values ["TransfPre"].replace(",", ".")
        TransfDia = values ["TransfDia"].replace(",", ".")

        listTrocoInicial.append(float(TrocoInicial))
        listEntradas.append(float(Entradas))
        listCartDEB.append(float(CartDEB))
        listDepoBAN.append(float(DepoBAN))
        listChequePRE.append(float(ChequePRE))
        listChequeDIA.append(float(ChequeDIA))
        listDepositoINTERNO.append(float(DepositoINTERNO))
        listOficioELE.append(float(OficioELE))
        listEinti.append(float(Einti))
        listErroFun.append(float(ErroFun))
        listErroSis.append(float(ErroSis))
        listdmp.append(float(dmp))
        listDebitoAuto.append(float(DebitoAuto))
        listNihil.append(float(Nihil))
        listCaixaFinal.append(float(CaixaFinal))
        listTransfQtT.append(float(TransfQtT))
        listTransfEspe.append(float(TransfEspe))
        listTransfPre.append(float(TransfPre))
        listTransfDia.append(float(TransfDia))
        janela.close()
        doc1.append(str(xdoc1))
        doc2.append(str(xdoc2))
        doc3.append(str(xdoc3))
        doc4.append(str(xdoc4))
        doc5.append(str(xdoc5))
        doc6.append(str(xdoc6))
        doc7.append(str(xdoc7))
        empresa.append(str(xempresa)) 
        por1.append(str(xpor1))
        por2.append(str(xpor2))

        

        return listTrocoInicial, listEntradas, listCartDEB, listDepoBAN, listChequePRE , listChequeDIA, listDepositoINTERNO, listOficioELE, listEinti, listErroFun, listErroSis, listdmp, listDebitoAuto, listNihil, listCaixaFinal, listTransfQtT, listTransfEspe, listTransfPre, listTransfDia,  doc1, doc2, doc3, doc4, doc5, doc6, doc7, empresa, por1, por2   
def Aviso():
        #layout
        sg.theme(tema)
        

        layout = [
            [sg.Text('INSIRA O NOME DO FUNCIONÁRIO E A DATA: ', size=(50,1))],
            [sg.Text('Funcionário: ', size=(20,1)), sg.Input(size=(10,1), key="user")],
            [sg.Text('Data: ', size=(20,1)), sg.Input(size=(10,1), key="data")],
            [sg.Button("Confirmar")],
        ]
        #janela
        janela = sg.Window("Alerta").layout(layout)
        events, values = janela.read()
        user.append(values["user"])
        data.append(values["data"])
       
        janela.Close()


        if events == "Confirmar":
            Saida()
def Saida():
        #layout
      
        color1 = "#9FB6CD"
        color2 = "#B9D3EE"
        color3 = "#9FB6CD"

        soma1 = float(sum(listCartDEB)) 
        soma2 = float(sum(listDepoBAN)) 
        soma3 = float(sum(listChequePRE)) 
        soma4 = float(sum(listChequeDIA))
        soma5 = float(sum(listDepositoINTERNO)) 
        soma6 = float(sum(listOficioELE)) 
        soma7 = float(sum(listEinti)) 
        soma8 = float(sum(listErroFun))
        soma9 = float(sum(listErroSis)) 
        soma10 = float(sum(listdmp))
        soma11 = float(sum(listDebitoAuto)) 
        soma12 = float(sum(listNihil))
        xsoma = soma1 + soma2 + soma3 + soma4 + soma5 + soma6 + soma7 + soma8 + soma9 + soma10 + soma11 + soma12
        especie1 = float(sum(listEntradas)) 
        especie2 = float(sum(listTransfPre))
        especie3 = float(sum(listTransfDia))
        xespecie = especie1 - xsoma - especie2 - especie3
        final1 = float(sum(listTransfEspe))
        final2 = float(sum(listTrocoInicial))
        xfinal = xespecie - final1 + final2
        diferenca1 = float(sum(listCaixaFinal))
        xdiferenca = diferenca1 - xfinal
        totaltransf1 = float(sum(listTransfEspe))
        totaltransf2 = float(sum(listTransfPre))
        totaltransf3 = float(sum(listTransfDia))
        
        xtotaltransf = totaltransf1 + totaltransf2 + totaltransf3

        if xdiferenca >= 0:
            cordiferenca = "#98FB98"
        else:
            cordiferenca = "#FF6347"
        
        sg.theme(tema)



        layTrocoInicial = [
            [sg.Text('TROCO INICIAL: R$ {:.2f}\n'.format(sum(listTrocoInicial)), background_color=color1)],
        ]
        layEntradas = [
            [sg.Text('ENTRADAS: R$ {:.2f}\n'.format(sum(listEntradas)), background_color=color1)],
        ]
        layCartDEB = [
            [sg.Text('1 - CARTÃO DE DÉBITO: R$ {:.2f} '.format(sum(listCartDEB)), background_color=color2)],
        ]
        layDepoBAN = [
            [sg.Text('2 - DEPÓSITO BANCÁRIO: R$ {:.2f} '.format(sum(listDepoBAN)), background_color=color2)],
        ]
        layChequePRE = [    
            [sg.Text('3 - CHEQUE PRÉ: R$ {:.2f} '.format(sum(listChequePRE)), background_color=color2)],
        ]
        layChequeDIA = [
            [sg.Text('4 - CHEQUE DIA: R$ {:.2f} '.format(sum(listChequeDIA)), background_color=color2)],
        ]
        layDepositoINTERNO = [
            [sg.Text('5 - DEPÓSITO INTERNO: R$ {:.2f} '.format(sum(listDepositoINTERNO)), background_color=color2)],  
        ]
        layOficioELE = [
            [sg.Text('6 - OFÍCIO ELETRÔNICO: R$ {:.2f}\n     Doc nº: {}'.format(sum(listOficioELE), "".join(doc1)), background_color=color2)],
        ]
        layEinti = [
            [sg.Text('7 - E-INTIMAÇÃO: R$ {:.2f}\n     Doc nº: {}'.format(sum(listEinti), "".join( doc2)), background_color=color2)],
        ]
        layErroFun = [
            [sg.Text('8 - ERRO DE FUNCIONÁRIO: R$ {:.2f}\n     Doc nº: {}'.format(sum(listErroFun), "".join(doc3)), background_color=color2)],
        ]
        layErroSis = [
            [sg.Text('9 - ERRO DO SISTEMA: R$ {:.2f}\n     Doc nº: {}'.format(sum(listErroSis), "".join(doc4)), background_color=color2)],
        ]
        laydmp = [
            [sg.Text('10 - DMP:  R$ {:.2f}\n     Doc nº: {}     Empresa: {}'.format(sum(listdmp), "".join(doc5), "".join(empresa)), background_color=color2)],
        ]
        layDebitoAuto = [
            [sg.Text('11 - DÉBITO AUTORIZADO: R$ {:.2f}\n      Doc nº: {}      Por: {}'.format(sum(listDebitoAuto), "".join(doc6), "".join(por1)), background_color=color2)],
        ] 
        layNihil = [    
            [sg.Text('12 - NIHIL: R$ {:.2f}\n     Doc nº: {}      Por: {}\n\n\n'.format(sum(listNihil), "".join(doc7), "".join(por2)), background_color=color2)],
        ]    
        laybotao = [ 
            [sg.Button("FINALIZAR", size=(15,1))],
        ]

        layTransf = [               
            [sg.Text('QUANTIDADES DE TRANSF: {:.2f}\n     ESPÉCIE: R${:.2f}\n     CHEQUE PRÉ: R${:.2f}\n     CHEQUE DIA: R${:.2f}\n     TOTAL DE TRANSF: R$ {:.2f}\n'.format(sum(listTransfQtT), sum(listTransfEspe), sum(listTransfPre), sum(listTransfDia), xtotaltransf), background_color=color3)],
        ]
        layespecie = [ 
            [sg.Text('ENTRADAS EM ESPÉCIE: R$ {:.2f}\n '.format(xespecie), background_color=color3)],
        ] 
        layCaixafinal = [ 
            [sg.Text('EM CAIXA: R$ {:.2f} '.format(sum(listCaixaFinal)), background_color=color3)],
        ]
        layFinal = [    
            [sg.Text("FECHAMENTO FINAL: R$ {:.2f} ".format(xfinal), background_color=color3)],
        ] 
        laydiferenca = [
            [sg.Text('DIFERENÇA EM CAIXA: R$ {:.2f} '.format(xdiferenca), background_color=cordiferenca)],
        ]

        layout = [
                [sg.Text("Funcionário: {} \nData: {}".format("".join(user), "".join(data)))],
                [sg.Frame("",layTrocoInicial, size=(300,43), background_color=color1), sg.Frame("",layOficioELE, size=(300,43), background_color=color2), sg.Frame("",layespecie, size=(300,43), background_color=color3)],
                [sg.Frame("",layEntradas, size=(300,43), background_color=color1), sg.Frame("",layEinti, size=(300,43), background_color=color2), sg.Frame("",layCaixafinal, size=(300,43), background_color=color3)],
                [sg.Frame("",layCartDEB, size=(300,43), background_color=color2), sg.Frame("",layErroFun, size=(300,43), background_color=color2), sg.Frame("",layFinal, size=(300,43), background_color=color3)],
                [sg.Frame("",layDepoBAN, size=(300,43), background_color=color2), sg.Frame("",layErroSis, size=(300,43), background_color=color2), sg.Frame("",laydiferenca, size=(300,43), background_color=cordiferenca)],
                [sg.Frame("",layChequePRE, size=(300,43), background_color=color2), sg.Frame("",laydmp, size=(300,43), background_color=color2)],
                [sg.Frame("",layChequeDIA, size=(300,43), background_color=color2), sg.Frame("",layDebitoAuto, size=(300,43), background_color=color2)],
                [sg.Frame("",layDepositoINTERNO, size=(300,43), background_color=color2), sg.Frame("",layNihil, size=(300,43), background_color=color2)],
                [sg.Frame("",layTransf, size=(300,100), background_color=color3), sg.Frame("",laybotao, size=(150,30))],
            ]       
                   

        #janela
        janela = sg.Window("  C A I X A    F E C H A D O  ").layout(layout)

        event, values = janela.read()

        janela.close
def CaixaAberto():
        #layout
        
        color1 = "#9FB6CD"
        color2 = "#B9D3EE"
        color3 = "#9FB6CD"

        sg.theme(tema)

        layTrocoInicial = [
            [sg.Text('TROCO INICIAL: R$ {:.2f}\n'.format(sum(listTrocoInicial)), background_color=color1)],
        ]
        layEntradas = [
            [sg.Text('ENTRADAS: R$ {:.2f}\n'.format(sum(listEntradas)), background_color=color1)],
        ]
        layCartDEB = [
            [sg.Text('1 - CARTÃO DE DÉBITO: R$ {:.2f} '.format(sum(listCartDEB)), background_color=color2)],
        ]
        layDepoBAN = [
            [sg.Text('2 - DEPÓSITO BANCÁRIO: R$ {:.2f} '.format(sum(listDepoBAN)), background_color=color2)],
        ]
        layChequePRE = [    
            [sg.Text('3 - CHEQUE PRÉ: R$ {:.2f} '.format(sum(listChequePRE)), background_color=color2)],
        ]
        layChequeDIA = [
            [sg.Text('4 - CHEQUE DIA: R$ {:.2f} '.format(sum(listChequeDIA)), background_color=color2)],
        ]
        layDepositoINTERNO = [
            [sg.Text('5 - DEPÓSITO INTERNO: R$ {:.2f} '.format(sum(listDepositoINTERNO)), background_color=color2)],  
        ]
        layOficioELE = [
            [sg.Text('6 - OFÍCIO ELETRÔNICO: R$ {:.2f}\n     Doc nº: {}'.format(sum(listOficioELE), "".join(doc1)), background_color=color2)],
        ]
        layEinti = [
            [sg.Text('7 - E-INTIMAÇÃO: R$ {:.2f}\n     Doc nº: {}'.format(sum(listEinti), "".join( doc2)), background_color=color2)],
        ]
        layErroFun = [
            [sg.Text('8 - ERRO DE FUNCIONÁRIO: R$ {:.2f}\n     Doc nº: {}'.format(sum(listErroFun), "".join(doc3)), background_color=color2)],
        ]
        layErroSis = [
            [sg.Text('9 - ERRO DO SISTEMA: R$ {:.2f}\n     Doc nº: {:}'.format(sum(listErroSis), "".join(doc4)), background_color=color2)],
        ]
        laydmp = [
            [sg.Text('10 - DMP:  R$ {:.2f}\n     Doc nº: {}     Empresa: {}'.format(sum(listdmp), "".join(doc5), "".join(empresa)), background_color=color2)],
        ]
        layDebitoAuto = [
            [sg.Text('11 - DÉBITO AUTORIZADO: R$ {:.2f}\n      Doc nº: {}      Por: {}'.format(sum(listDebitoAuto), "".join(doc6), "".join(por1)), background_color=color2)],
        ] 
        layNihil = [    
            [sg.Text('12 - NIHIL: R$ {:.2f}\n     Doc nº: {}      Por: {}\n\n\n'.format(sum(listNihil), "".join(doc7), "".join(por2)), background_color=color2)],
        ]    
        laybotao = [ 
            [sg.Button("ACRESCENTAR PAGAMENTO", size=(15,2))],
        ]
        laybotao2 = [
            [sg.Button("FECHAR \nCAIXA", size=(15,2))],
        ]

        layTransf = [               
            [sg.Text('QUANTIDADES DE TRANSF: {:.2f}\n     ESPÉCIE: R${:.2f}\n     CHEQUE PRÉ: R${:.2f}\n     CHEQUE DIA: R${:.2f}\n     TOTAL DE TRANSF: R$ {:.2f}\n'.format(sum(listTransfQtT), sum(listTransfEspe), sum(listTransfPre), sum(listTransfDia), sum(totaltransf)), background_color=color3)],
        ]
        
        layCaixafinal = [ 
            [sg.Text('EM CAIXA: R$ {:.2f} '.format(sum(listCaixaFinal)), background_color=color3)],
        ]

        layout = [
                [sg.Frame("",layTrocoInicial, size=(300,43), background_color=color1), sg.Frame("",layOficioELE, size=(300,43), background_color=color2), sg.Frame("",layCaixafinal, size=(300,43), background_color=color3)],
                [sg.Frame("",layEntradas, size=(300,43), background_color=color1), sg.Frame("",layEinti, size=(300,43), background_color=color2)],
                [sg.Frame("",layCartDEB, size=(300,43), background_color=color2), sg.Frame("",layErroFun, size=(300,43), background_color=color2)],
                [sg.Frame("",layDepoBAN, size=(300,43), background_color=color2), sg.Frame("",layErroSis, size=(300,43), background_color=color2)], 
                [sg.Frame("",layChequePRE, size=(300,43), background_color=color2), sg.Frame("",laydmp, size=(300,43), background_color=color2)],
                [sg.Frame("",layChequeDIA, size=(300,43), background_color=color2), sg.Frame("",layDebitoAuto, size=(300,43), background_color=color2)],
                [sg.Frame("",layDepositoINTERNO, size=(300,43), background_color=color2), sg.Frame("",layNihil, size=(300,43), background_color=color2)],
                [sg.Frame("",layTransf, size=(300,100), background_color=color3), sg.Frame("",laybotao, size=(150,50)), sg.Frame("",laybotao2, size=(150,50))],
            ]       

        #janela
        janela = sg.Window("  C A I X A   A B E R T O   ").layout(layout)

        event, values = janela.read()
        if event == "ACRESCENTAR PAGAMENTO":
            janela.close()
            loop()
        elif event == "FECHAR \nCAIXA":
            janela.close()
            Aviso()        
def listas():
        listTrocoInicial = []
        listEntradas = []
        listCartDEB = []
        listDepoBAN = []
        listChequePRE = []
        listChequeDIA = []
        listDepositoINTERNO = []
        listOficioELE = []
        listEinti = []
        listErroFun = []
        listErroSis = []
        listdmp = []
        listDebitoAuto = []
        listNihil = []
        listCaixaFinal = []
        listTransfQtT = []
        listTransfEspe = []
        listTransfPre = []
        listTransfDia = []
        doc1 = [] 
        doc2 = []
        doc3 = []
        doc4 = [] 
        doc5 = []
        doc6 = [] 
        doc7 = []
        empresa = [] 
        por1 = []
        por2 = []
        soma = []
        especie = []
        final = []
        diferenca = []
        totaltransf = []
        aviso = []
        user = []
        data = []

        xdiferenca = sum(diferenca)
                
        if xdiferenca >= 0:
            aviso = "Sobrou: R$"
        else:
            aviso = "Faltou: R$"


        return listTrocoInicial, listEntradas, listCartDEB, listDepoBAN, listChequePRE , listChequeDIA, listDepositoINTERNO, listOficioELE, listEinti, listErroFun, listErroSis, listdmp, listDebitoAuto, listNihil, listCaixaFinal, listTransfQtT, listTransfEspe, listTransfPre, listTransfDia, doc1, doc2, doc3, doc4, doc5, doc6, doc7, empresa, por1, por2, soma, especie, final, diferenca, totaltransf, aviso, user, data
def Funcionario():
        if fun == "Renata":
            wb = load_workbook(filename= "CAIXA 1.xlsx")
            nomearquivo = "CAIXA 1.xlsx"
        elif fun == "Suellen":
            wb = load_workbook(filename= "CAIXA 2.xlsx")
            nomearquivo = "CAIXA 2.xlsx"
        elif fun == "Moises":
            wb = load_workbook(filename= "CAIXA 3.xlsx")
            nomearquivo = "CAIXA 3.xlsx"
        elif fun == "Endrielly":
            wb = load_workbook(filename= "CAIXA 4.xlsx")
            nomearquivo = "CAIXA 4.xlsx"
        return nomearquivo, wb
def Dia():
        if dia == "01":
            sh = wb["DIA 01"]
        elif dia == "02":
            sh = wb["DIA 02"]
        elif dia == "03":
            sh = wb["DIA 03"]
        elif dia == "04":
            sh = wb["DIA 04"]
        elif dia == "05":
            sh = wb["DIA 05"]
        elif dia == "06":
            sh = wb["DIA 06"]
        elif dia == "07":
            sh = wb["DIA 07"]
        elif dia == "08":
            sh = wb["DIA 08"]
        elif dia == "09":
            sh = wb["DIA 09"]
        elif dia == "10":
            sh = wb["DIA 10"]
        elif dia == "11":
            sh = wb["DIA 11"]
        elif dia == "12":
            sh = wb["DIA 12"]
        elif dia == "13":
            sh = wb["DIA 13"]
        elif dia == "14":
            sh = wb["DIA 14"]
        elif dia == "15":
            sh = wb["DIA 15"]
        elif dia == "16":
            sh = wb["DIA 16"]
        elif dia == "17":
            sh = wb["DIA 17"]
        elif dia == "18":
            sh = wb["DIA 18"]
        elif dia == "19":
            sh = wb["DIA 19"]
        elif dia == "20":
            sh = wb["DIA 20"]
        elif dia == "21":
            sh = wb["DIA 21"]
        elif dia == "22":
            sh = wb["DIA 22"]
        elif dia == "23":
            sh = wb["DIA 23"]
        elif dia == "24":
            sh = wb["DIA 24"]
        elif dia == "25":
            sh = wb["DIA 25"]
        elif dia == "26":
            sh = wb["DIA 26"]
        elif dia == "27":
            sh = wb["DIA 27"]
        elif dia == "28":
            sh = wb["DIA 28"]
        elif dia == "29":
            sh = wb["DIA 29"]
        elif dia == "30":
            sh = wb["DIA 30"]
        elif dia == "31":
            sh = wb["DIA 31"]
        return sh
def loop(): 
    Entrada()
    CaixaAberto()
    
fun, dia = Escolha()
nomearquivo, wb = Funcionario()
sh = Dia()

listTrocoInicial, listEntradas, listCartDEB, listDepoBAN, listChequePRE , listChequeDIA, listDepositoINTERNO, listOficioELE, listEinti, listErroFun, listErroSis, listdmp, listDebitoAuto, listNihil, listCaixaFinal, listTransfQtT, listTransfEspe, listTransfPre, listTransfDia, doc1, doc2, doc3, doc4, doc5, doc6, doc7, empresa, por1, por2, soma, especie, final, diferenca, totaltransf, aviso, user, data = listas()
  
loop()

#alocação no excel

sh["D5"] = float(sum(listTrocoInicial))
sh["D7"] = float(sum(listEntradas))
sh["D9"] = float(sum(listCartDEB))
sh["D10"] = float(sum(listDepoBAN))
sh["D11"] = float(sum(listChequePRE))
sh["D12"] = float(sum(listChequeDIA))
sh["D13"] = float(sum(listDepositoINTERNO))
sh["D14"] = float(sum(listOficioELE))
sh["D15"] = float(sum(listEinti))
sh["D16"] = float(sum(listErroFun))
sh["D17"] = float(sum(listErroSis))
sh["D18"] = float(sum(listdmp))
sh["D19"] =  float(sum(listDebitoAuto))
sh["D20"] = float(sum(listNihil))
sh["I6"] = float(sum(listCaixaFinal))
sh["F14"] = ",".join(doc1)
sh["F15"] = ",".join(doc2)
sh["F16"] = ",".join(doc3)
sh["F17"] = ",".join(doc4)
sh["F18"] = ",".join(doc5)
sh["F19"] = ",".join(doc6)
sh["F20"] = ",".join(doc7)
sh["I18"] = ",".join(empresa)
sh["I19"] = ",".join(por1)
sh["I20"] = ",".join(por2)
sh["D23"] = float(sum(listTransfQtT))
sh["F23"] = float(sum(listTransfEspe))
sh["I23"] = float(sum(listTransfPre))
sh["I24"] = float(sum(listTransfDia))

wb.save(filename = nomearquivo)